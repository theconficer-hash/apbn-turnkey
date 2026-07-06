"""Ekspor hasil simulasi APBN Turnkey ke workbook Excel (openpyxl)."""
import io

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Palet selaras UI (laut & mangrove)
HEADER_FILL = PatternFill("solid", fgColor="1F5E8C")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="EAF3FA")
TOTAL_FONT = Font(bold=True, color="123F63")
TITLE_FONT = Font(bold=True, size=14, color="123F63")
THIN = Side(style="thin", color="D6E7F4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FMT_RP = '#,##0'


def _tulis_sheet(ws, judul, header, rows, total_row=None):
    ws.append([judul])
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.append([])

    ws.append(header)
    hr = ws.max_row
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=hr, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for row in rows:
        ws.append(row)
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.border = BORDER
            if isinstance(cell.value, (int, float)) and c > 1:
                cell.number_format = FMT_RP

    if total_row:
        ws.append(total_row)
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = BORDER
            if isinstance(cell.value, (int, float)) and c > 1:
                cell.number_format = FMT_RP

    # Lebar kolom
    ws.column_dimensions["A"].width = 22
    for c in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20
    ws.freeze_panes = ws.cell(row=hr + 1, column=2)


def build_excel(hasil: dict, asumsi: dict) -> bytes:
    years = hasil["tahun_aktif"]
    per_tahap = hasil["per_tahap"]
    m = hasil["metrics"]

    wb = Workbook()

    # ── Sheet 1: Ringkasan ──
    ws = wb.active
    ws.title = "Ringkasan"
    ws.append(["Simulator APBN Turnkey — Ringkasan Hasil"])
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.append([])
    info = [
        ("Cost of Debt", f"{asumsi['cost_of_debt'] * 100:.2f}%"),
        ("Tenor Pinjaman", f"{asumsi['tenor']} tahun"),
        ("OPEX Rate", f"{asumsi['opex_rate'] * 100:.2f}% dari CAPEX/tahun"),
        ("Masa Operasi", f"{asumsi['masa_operasi']} tahun"),
        ("Horizon Simulasi", f"{years[0]}–{years[-1]}" if years else "-"),
        (),
        ("Total CAPEX (Rp)", m["total_capex"]),
        ("Beban Puncak per Tahun (Rp)", m["beban_puncak"]),
        ("Tahun Puncak", m["tahun_puncak"]),
        ("Total Kumulatif Beban (Rp)", m["total_kumulatif"]),
        ("  — Total Anuitas (Rp)", m["total_anuitas_kumulatif"]),
        ("  — Total OPEX (Rp)", m["total_opex_kumulatif"]),
        ("Jumlah Tahap Aktif", m["jumlah_tahap"]),
    ]
    for row in info:
        ws.append(list(row))
        if row and len(row) > 1:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            v = ws.cell(row=ws.max_row, column=2)
            if isinstance(v.value, (int, float)):
                v.number_format = FMT_RP
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 26

    # ── Sheet 2: Beban per Tahun ──
    _tulis_sheet(
        wb.create_sheet("Beban per Tahun"),
        "Total Beban per Tahun (Rp)",
        ["Tahun", "Total Anuitas", "Total OPEX", "Total Beban"],
        [
            [y,
             round(hasil["total_anuitas"].get(y, 0)),
             round(hasil["total_opex"].get(y, 0)),
             round(hasil["total_beban"].get(y, 0))]
            for y in years
        ],
        total_row=[
            "TOTAL",
            round(m["total_anuitas_kumulatif"]),
            round(m["total_opex_kumulatif"]),
            round(m["total_kumulatif"]),
        ],
    )

    # ── Sheet 3: Ringkasan per Tahap ──
    _tulis_sheet(
        wb.create_sheet("Ringkasan per Tahap"),
        "Ringkasan Keuangan per Tahap (Rp)",
        ["Tahap", "CAPEX", "Anuitas per Tahun", "OPEX per Tahun", "Beban per Tahun",
         "Total Cicilan", "Konstruksi Mulai", "Konstruksi Selesai",
         "Cicilan Mulai", "Cicilan Selesai"],
        [
            [t["nama"], round(t["capex"]), round(t["anuitas_tahunan"]),
             round(t["opex_tahunan"]), round(t["total_beban_tahunan"]),
             round(t["total_cicilan"]), t["mulai_konstruksi"], t["selesai_konstruksi"],
             t["cicilan_mulai"], t["cicilan_selesai"]]
            for t in per_tahap
        ],
    )

    # ── Sheet 4: Beban per Tahap per Tahun ──
    _tulis_sheet(
        wb.create_sheet("Beban per Tahap per Tahun"),
        "Beban per Tahap per Tahun — Anuitas + OPEX (Rp)",
        ["Tahun", *[t["nama"] for t in per_tahap], "Total"],
        [
            [y,
             *[round(t["anuitas_per_tahun"].get(y, 0) + t["opex_per_tahun"].get(y, 0))
               for t in per_tahap],
             round(hasil["total_beban"].get(y, 0))]
            for y in years
        ],
    )

    # ── Sheet 5: Grafik (chart Excel native, data dalam Rp Triliun) ──
    ws = wb.create_sheet("Grafik")
    ws.append(["Data Grafik (Rp Triliun)"])
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.append([])

    # Tabel data: Tahun | tahap1..N | Total Anuitas | Total OPEX | Total Beban
    nama_tahap = [t["nama"] for t in per_tahap]
    header = ["Tahun", *nama_tahap, "Total Anuitas", "Total OPEX", "Total Beban"]
    ws.append(header)
    hr = ws.max_row  # baris header data
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=hr, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER

    T = 1e12
    for y in years:
        ws.append([
            y,
            *[round((t["anuitas_per_tahun"].get(y, 0) + t["opex_per_tahun"].get(y, 0)) / T, 4)
              for t in per_tahap],
            round(hasil["total_anuitas"].get(y, 0) / T, 4),
            round(hasil["total_opex"].get(y, 0) / T, 4),
            round(hasil["total_beban"].get(y, 0) / T, 4),
        ])
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.border = BORDER
            if c > 1:
                cell.number_format = "0.00"

    baris_awal = hr + 1
    baris_akhir = ws.max_row
    n_tahap = len(nama_tahap)
    kolom_tahun = Reference(ws, min_col=1, min_row=baris_awal, max_row=baris_akhir)

    # Grafik 1 — Beban per Tahap (stacked bar), warna selaras UI
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "stacked"
    bar.overlap = 100
    bar.title = "Beban per Tahap (Rp Triliun / tahun)"
    bar.y_axis.title = "Rp Triliun"
    bar.x_axis.title = "Tahun"
    data = Reference(ws, min_col=2, max_col=1 + n_tahap,
                     min_row=hr, max_row=baris_akhir)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(kolom_tahun)
    warna_ui = ["1F5E8C", "1B7A5A", "B45309", "7C3AED", "0E7490", "BE123C"]
    for i, s in enumerate(bar.series):
        s.graphicalProperties.solidFill = warna_ui[i % len(warna_ui)]
    bar.width = 30
    bar.height = 12
    ws.add_chart(bar, f"A{baris_akhir + 3}")

    # Grafik 2 — Anuitas vs OPEX vs Total (line)
    line = LineChart()
    line.title = "Anuitas vs OPEX Total (Rp Triliun / tahun)"
    line.y_axis.title = "Rp Triliun"
    line.x_axis.title = "Tahun"
    data = Reference(ws, min_col=2 + n_tahap, max_col=4 + n_tahap,
                     min_row=hr, max_row=baris_akhir)
    line.add_data(data, titles_from_data=True)
    line.set_categories(kolom_tahun)
    for s, warna in zip(line.series, ["1F5E8C", "1B7A5A", "DC2626"]):
        s.graphicalProperties.line.solidFill = warna
        s.graphicalProperties.line.width = 25000  # ~2pt (EMU)
        s.smooth = False
    line.width = 30
    line.height = 12
    ws.add_chart(line, f"A{baris_akhir + 28}")

    ws.column_dimensions["A"].width = 10
    for c in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = ws.cell(row=baris_awal, column=2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
